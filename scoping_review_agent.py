#!/usr/bin/env python3
"""
Scoping Review Data Extraction Agent
Automatically extracts data from academic PDFs in Google Drive and populates Excel sheet for coastal health scoping review.
"""

import os
import json
import logging
import sys
import re
import io
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
from dotenv import load_dotenv

import pdfplumber
from pypdf import PdfReader
import openpyxl
from openpyxl.utils import get_column_letter

# Google Drive API imports
from google.auth.transport.requests import Request
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# Load environment variables
load_dotenv()

# ============================================================================
# CONFIGURATION
# ============================================================================

# Google Drive configuration
GOOGLE_DRIVE_FOLDER_ID = os.getenv("GOOGLE_DRIVE_FOLDER_ID")
GOOGLE_CLOUD_CREDENTIALS_JSON = os.getenv("GOOGLE_CLOUD_CREDENTIALS")

# Your Excel file
EXCEL_FILE = "Data Extraction Form_1.xlsx"

# API and tracking
PROCESSED_FILES_LOG = "processed_files.json"
LOG_FILE = "extraction_agent.log"

# ============================================================================
# LOGGING SETUP
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# ============================================================================
# LOCAL EXTRACTION PATTERNS & UTILITIES
# ============================================================================

STUDY_DESIGNS = [
    "RCT", "randomized controlled trial", "cohort", "case control", "cross-sectional",
    "qualitative", "mixed methods", "longitudinal", "retrospective", "prospective",
    "phenomenological", "grounded theory", "ethnographic", "experimental", "quasi-experimental"
]

EVIDENCE_TYPES = [
    "Primary research", "Epidemiology", "Evidence synthesis", "Conference abstract",
    "Editorial", "Discussion article"
]

COUNTRIES = {
    "UK": ["UK", "United Kingdom", "Scotland", "England", "Wales", "Northern Ireland"],
    "USA": ["USA", "United States", "US"],
    "Australia": ["Australia", "Australian"],
    "Canada": ["Canada", "Canadian"],
    "New Zealand": ["New Zealand"],
}

# ============================================================================
# GOOGLE DRIVE API FUNCTIONS
# ============================================================================

def get_drive_service():
    """Initialize and return Google Drive API service."""
    if not GOOGLE_CLOUD_CREDENTIALS_JSON:
        logger.error("GOOGLE_CLOUD_CREDENTIALS not set")
        return None

    try:
        creds_dict = json.loads(GOOGLE_CLOUD_CREDENTIALS_JSON)
        credentials = Credentials.from_service_account_info(creds_dict)
        service = build('drive', 'v3', credentials=credentials)
        return service
    except Exception as e:
        logger.error(f"Failed to initialize Google Drive service: {e}")
        return None

def get_pdfs_from_drive() -> List[Tuple[str, str]]:
    """Get list of PDFs from Google Drive folder. Returns list of (file_id, file_name)."""
    if not GOOGLE_DRIVE_FOLDER_ID:
        logger.error("GOOGLE_DRIVE_FOLDER_ID not set")
        return []

    service = get_drive_service()
    if not service:
        return []

    try:
        query = f"'{GOOGLE_DRIVE_FOLDER_ID}' in parents and mimeType='application/pdf' and trashed=false"
        results = service.files().list(
            q=query,
            spaces='drive',
            fields='files(id, name)',
            pageSize=100
        ).execute()

        files = results.get('files', [])
        return [(f['id'], f['name']) for f in files]
    except Exception as e:
        logger.error(f"Failed to get PDFs from Google Drive: {e}")
        return []

def download_pdf_from_drive(file_id: str) -> Optional[bytes]:
    """Download PDF file content from Google Drive as bytes."""
    service = get_drive_service()
    if not service:
        return None

    try:
        request = service.files().get_media(fileId=file_id)
        file_content = io.BytesIO()
        downloader = MediaIoBaseDownload(file_content, request)

        done = False
        while not done:
            status, done = downloader.next_chunk()

        return file_content.getvalue()
    except Exception as e:
        logger.error(f"Failed to download PDF from Google Drive: {e}")
        return None

def extract_year_from_text(text: str) -> str:
    """Extract publication year from text."""
    # Look for 4-digit years between 1900 and 2100
    years = re.findall(r'\b(19\d{2}|20\d{2})\b', text[:1000])
    if years:
        return str(max(int(y) for y in years))
    return "N/A"

def extract_authors_from_text(text: str) -> str:
    """Extract authors - usually appear after title (lines 2-15)."""
    lines = text.split('\n')

    # Look for author names in lines 2-15 (after title, before body text)
    for i in range(2, min(16, len(lines))):
        line = lines[i].strip()

        # Author line characteristics:
        # - Not too long (usually < 120 chars)
        # - Contains capital letters and at least one space
        # - Not a sentence, not title continuation
        # - Not a section header (not all caps)
        if (8 < len(line) < 120 and
            line[0].isupper() and
            not line.isupper() and
            ' ' in line and
            not any(x in line[:30] for x in ['as ', 'in ', 'the ', 'This', 'Abstract', 'Introduction', 'Working on', 'Exploring the'])):

            # Extract just the author name (remove degrees/affiliation)
            author = line.split(' BSc')[0].split(' PhD')[0].split(' MSc')[0].split(' MD')[0].split(' RN')[0].split(' DVM')[0]
            if ' is ' in author:
                author = author.split(' is ')[0]

            author = author.strip()

            # Validate (has capitals)
            if len(author) > 4 and re.search(r'[A-Z].*[A-Z]', author):
                return author

    return "N/A"

def extract_title_from_metadata(pdf_path: str, text: str) -> str:
    """Extract title from PDF metadata or first lines of text."""
    try:
        reader = PdfReader(pdf_path)
        if reader.metadata and reader.metadata.get('/Title'):
            return reader.metadata['/Title']
    except:
        pass

    # Fallback: extract from first 200 chars (often the title)
    if text:
        lines = [l.strip() for l in text.split('\n')[:5] if l.strip()]
        if lines:
            # Return first substantial line (not author names)
            for line in lines:
                if len(line) > 10 and len(line) < 300 and not any(x in line for x in ['BSc', 'PhD', 'MSc', 'Prof']):
                    return line
            return lines[0] if lines else "N/A"

    return "N/A"

def extract_country(text: str) -> str:
    """Extract country from text."""
    for country, keywords in COUNTRIES.items():
        for keyword in keywords:
            if keyword.lower() in text.lower():
                return country
    # Look for common country mentions
    if "Australia" in text or "Australian" in text:
        return "Australia"
    if "NHS" in text or "NICE" in text:
        return "UK"
    return "N/A"

def find_section(text: str, section_name: str, max_chars: int = 2000, variations: list = None) -> str:
    """Extract content from a specific section (Abstract, Methods, Results, etc).
    Tries multiple section name variations."""
    if variations is None:
        variations = [section_name]

    lines = text.split('\n')
    start_idx = -1

    # Find section start - try all variations
    for i, line in enumerate(lines):
        for variant in variations:
            if variant.lower() in line.lower():
                start_idx = i + 1
                break
        if start_idx != -1:
            break

    if start_idx == -1:
        return "N/A"

    # Extract until next section or max chars
    section_text = []
    char_count = 0
    for line in lines[start_idx:]:
        # Stop at next section header (all caps, multiple words)
        if len(line.strip()) > 3 and line.strip().isupper() and len(line.strip().split()) >= 1:
            break
        section_text.append(line)
        char_count += len(line)
        if char_count > max_chars:
            break

    result = ' '.join(section_text).strip()
    return result if result else "N/A"

def classify_study_design(text: str) -> str:
    """Classify study design from text."""
    text_lower = text.lower()
    for design in STUDY_DESIGNS:
        if design.lower() in text_lower:
            return design.capitalize()
    return "N/A"

def extract_coastal_definition(text: str) -> str:
    """Extract coastal area definition from text."""
    coastal_patterns = [
        r'coastal(?:\s+area)?.*?(?:defined|defined as|km|mile|kilometer)',
        r'coastline.*?(?:within|distance|km)',
        r'definition.*?coastal'
    ]

    for pattern in coastal_patterns:
        matches = re.finditer(pattern, text, re.IGNORECASE | re.DOTALL)
        for match in matches:
            return match.group(0)[:300]

    return "N/A"

def extract_data_locally(paper_text: str, pdf_path: str) -> Optional[Dict[str, str]]:
    """Extract data using local NLP and pattern matching (no API)."""
    try:
        # Extract sections with multiple name variations
        abstract = find_section(paper_text, "abstract", 1500, ["abstract", "aims", "objective"])
        methods = find_section(paper_text, "method", 2000, ["method", "methods", "methodology", "procedure"])
        results = find_section(paper_text, "result", 2000, ["result", "results", "findings"])
        discussion = find_section(paper_text, "discussion", 1500, ["discussion", "implications"])

        # Extract title (now with text fallback)
        title = extract_title_from_metadata(pdf_path, paper_text)

        # Extract publication venue (journal name, organization, etc.)
        pub_venue = "N/A"
        # Look for patterns like "Published in", "Journal of", "Journal:", etc.
        venue_match = re.search(r'(?:published in|journal of|journal:|journal\s+name)[:\s]+([^.\n]+)', paper_text, re.IGNORECASE)
        if venue_match:
            pub_venue = venue_match.group(1).strip()[:100]
        else:
            # Try to find in the first 500 chars (sometimes in subtitle)
            first_section = paper_text[:500]
            if 'Journal' in first_section:
                j_match = re.search(r'([\w\s&]+Journal[\w\s&]+)', first_section)
                if j_match:
                    pub_venue = j_match.group(1).strip()[:100]

        # Extract population/sample size - more flexible pattern
        sample_size = "N/A"
        sample_match = re.search(r'n\s*=\s*(\d+)|sample\s*(?:size)?:?\s*(\d+)|N\s*=\s*(\d+)', paper_text, re.IGNORECASE)
        if sample_match:
            n_value = sample_match.group(1) or sample_match.group(2) or sample_match.group(3)
            sample_size = f"n={n_value}"

        # Extract research questions/aims text
        research_q = "N/A"
        if "research question" in paper_text.lower():
            research_q = find_section(paper_text, "research question", 300)[:200]
        elif "aims" in paper_text.lower():
            aims_section = find_section(paper_text, "aims", 500)
            if aims_section != "N/A":
                research_q = aims_section[:200]

        # Funding
        funding = "N/A"
        if re.search(r'fund|grant|support', paper_text, re.IGNORECASE):
            funding = find_section(paper_text, "funding", 300, ["funding", "acknowledgment", "acknowledgements"])
            if funding == "N/A":
                funding_match = re.search(r'(?:funded|supported|grant).{0,100}', paper_text, re.IGNORECASE)
                if funding_match:
                    funding = funding_match.group(0)

        # Affiliation
        affiliation = "N/A"
        aff_match = re.search(r'(?:affiliation|department|university|institute|school).{0,100}', paper_text, re.IGNORECASE)
        if aff_match:
            affiliation = aff_match.group(0)

        extracted_data = {
            "s_no": "TBD",
            "authors": extract_authors_from_text(paper_text),
            "country_location": extract_country(paper_text),
            "title": title,
            "type_of_publication": "Journal article" if "journal" in paper_text.lower() else "N/A",
            "publication_venue": pub_venue,
            "publication_year": extract_year_from_text(paper_text),
            "type_of_evidence_source": "Primary research" if any(x in paper_text.lower() for x in ["study", "research", "audit"]) else "N/A",
            "aim_purpose": abstract[:400] if abstract != "N/A" else "N/A",
            "research_questions_hypothesis": research_q,
            "study_design": classify_study_design(methods if methods != "N/A" else paper_text),
            "methodology": methods[:600] if methods != "N/A" else "N/A",
            "population_sample_size": sample_size,
            "setting_context": find_section(paper_text, "setting", 500, ["setting", "context", "location"]),
            "date_location_data_collection": re.findall(r'\d{4}[-/]\d{2}|20\d{2}', paper_text)[-1] if re.findall(r'\d{4}[-/]\d{2}|20\d{2}', paper_text) else "N/A",
            "definition_coastal_areas": extract_coastal_definition(paper_text),
            "interventions_programmes": find_section(paper_text, "intervention", 800, ["intervention", "programme", "program", "treatment"]),
            "outcomes": find_section(paper_text, "outcome", 800, ["outcome", "results", "findings"]),
            "key_findings": results[:500] if results != "N/A" else "N/A",
            "knowledge_gaps": discussion[:500] if discussion != "N/A" else "N/A",
            "barriers_facilitators_limitations": find_section(paper_text, "limitation", 600, ["limitation", "barrier", "challenge", "facilitator"]),
            "institutional_affiliation": affiliation,
            "funder_funding": funding,
            "data_accessibility": "",
        }

        return extracted_data
    except Exception as e:
        logger.error(f"Local extraction failed: {e}")
        return None

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def load_processed_files() -> set:
    """Load the list of already-processed PDF filenames."""
    if os.path.exists(PROCESSED_FILES_LOG):
        try:
            with open(PROCESSED_FILES_LOG, 'r') as f:
                data = json.load(f)
                return set(data.get("processed_files", []))
        except json.JSONDecodeError:
            logger.warning(f"Could not read {PROCESSED_FILES_LOG}, starting fresh")
            return set()
    return set()


def save_processed_files(processed: set) -> None:
    """Save the list of processed filenames."""
    with open(PROCESSED_FILES_LOG, 'w') as f:
        json.dump({"processed_files": sorted(list(processed))}, f, indent=2)


def extract_text_from_pdf_bytes(pdf_bytes: bytes, filename: str = "PDF") -> Optional[str]:
    """Extract text from PDF bytes using pdfplumber, fallback to pypdf."""
    try:
        # Try pdfplumber first
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            if text.strip():
                return text
    except Exception as e:
        logger.debug(f"pdfplumber failed for {filename}: {e}, trying pypdf")

    # Fallback to pypdf
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        if text.strip():
            return text
    except Exception as e:
        logger.error(f"Failed to extract text from {filename} using both methods: {e}")

    return None

def extract_text_from_pdf(pdf_path: str) -> Optional[str]:
    """Extract text from a local PDF file using pdfplumber, fallback to pypdf."""
    try:
        # Try pdfplumber first
        with pdfplumber.open(pdf_path) as pdf:
            text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            if text.strip():
                return text
    except Exception as e:
        logger.debug(f"pdfplumber failed for {pdf_path}: {e}, trying pypdf")

    # Fallback to pypdf
    try:
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        if text.strip():
            return text
    except Exception as e:
        logger.error(f"Failed to extract text from {pdf_path} using both methods: {e}")

    return None


def extract_data_with_claude(paper_text: str) -> Optional[Dict[str, str]]:
    """Send paper text to Claude API and get structured extraction."""
    try:
        client = Anthropic()

        # Check for API key
        if not os.getenv("ANTHROPIC_API_KEY"):
            logger.error("ANTHROPIC_API_KEY environment variable not set")
            return None

        # Prepare the full prompt
        full_prompt = EXTRACTION_PROMPT.format(paper_text=paper_text)

        # Call Claude API
        message = client.messages.create(
            model="claude-3-5-sonnet-20241022",
            max_tokens=2000,
            messages=[
                {"role": "user", "content": full_prompt}
            ]
        )

        # Extract and parse JSON response
        response_text = message.content[0].text

        # Try to parse JSON from response
        try:
            extracted_data = json.loads(response_text)
            return extracted_data
        except json.JSONDecodeError:
            logger.error(f"Claude returned invalid JSON: {response_text[:200]}")
            return None

    except Exception as e:
        logger.error(f"Claude API call failed: {e}")
        return None


def get_last_row_with_data(ws) -> int:
    """Find the last row with data in the worksheet (consecutive rows, no alternating structure)."""
    last_row = 2
    for row in range(3, ws.max_row + 1):
        # Check if column A has a numeric serial number
        col_a = ws[f"A{row}"].value
        if col_a is not None:
            try:
                int(str(col_a).strip())
                last_row = row
            except (ValueError, TypeError):
                pass
    return last_row


def get_next_serial_number(ws) -> int:
    """Get the next serial number to use."""
    last_row = get_last_row_with_data(ws)
    if last_row == 2:
        return 1
    try:
        last_sno = int(ws[f"A{last_row}"].value)
        return last_sno + 1
    except (ValueError, TypeError):
        return last_row - 1  # Fallback


def write_extraction_to_excel(extracted_data: Dict[str, str], pdf_filename: str) -> bool:
    """Write extracted data to Excel sheet.
    Appends to next ODD row (data rows) after existing data."""
    try:
        # Load workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        # Find next empty row (consecutive rows, no alternating structure)
        last_row = get_last_row_with_data(ws)
        new_row = last_row + 1

        # Get next serial number
        next_sno = get_next_serial_number(ws)

        # Get current date and time for extraction timestamp
        extraction_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Define the mapping from extracted data keys to Excel columns
        # Order: A-X (24 fields) + Y (extraction date/time)
        field_mapping = {
            "s_no": ("A", str(next_sno)),
            "authors": ("B", extracted_data.get("authors", "N/A")),
            "country_location": ("C", extracted_data.get("country_location", "N/A")),
            "title": ("D", extracted_data.get("title", "N/A")),
            "type_of_publication": ("E", extracted_data.get("type_of_publication", "N/A")),
            "publication_venue": ("F", extracted_data.get("publication_venue", "N/A")),  # New: where published
            "publication_year": ("G", extracted_data.get("publication_year", "N/A")),
            "type_of_evidence_source": ("H", extracted_data.get("type_of_evidence_source", "N/A")),
            "aim_purpose": ("I", extracted_data.get("aim_purpose", "N/A")),
            "research_questions_hypothesis": ("J", extracted_data.get("research_questions_hypothesis", "N/A")),
            "study_design": ("K", extracted_data.get("study_design", "N/A")),
            "methodology": ("L", extracted_data.get("methodology", "N/A")),
            "population_sample_size": ("M", extracted_data.get("population_sample_size", "N/A")),
            "setting_context": ("N", extracted_data.get("setting_context", "N/A")),
            "date_location_data_collection": ("O", extracted_data.get("date_location_data_collection", "N/A")),
            "definition_coastal_areas": ("P", extracted_data.get("definition_coastal_areas", "N/A")),
            "interventions_programmes": ("Q", extracted_data.get("interventions_programmes", "N/A")),
            "outcomes": ("R", extracted_data.get("outcomes", "N/A")),
            "key_findings": ("S", extracted_data.get("key_findings", "N/A")),
            "knowledge_gaps": ("T", extracted_data.get("knowledge_gaps", "N/A")),
            "barriers_facilitators_limitations": ("U", extracted_data.get("barriers_facilitators_limitations", "N/A")),
            "institutional_affiliation": ("V", extracted_data.get("institutional_affiliation", "N/A")),
            "funder_funding": ("W", extracted_data.get("funder_funding", "N/A")),
            "data_accessibility": ("X", ""),  # Leave blank for manual entry
            "extraction_date_time": ("Y", extraction_timestamp),
        }

        # Write data to cells
        for key, (col, value) in field_mapping.items():
            cell = ws[f"{col}{new_row}"]
            cell.value = value
            # Wrap text for long fields
            if key in ["methodology", "key_findings", "barriers_facilitators_limitations"]:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # Save workbook
        wb.save(EXCEL_FILE)
        logger.info(f"Data written to Excel row {new_row}: {pdf_filename}")
        return True

    except Exception as e:
        logger.error(f"Failed to write data to Excel: {e}")
        return False


def find_new_pdfs() -> List[str]:
    """Find all new PDFs in the papers folder (for backward compatibility)."""
    papers_path = Path(PAPERS_FOLDER)

    if not papers_path.exists():
        logger.warning(f"Papers folder does not exist: {PAPERS_FOLDER}")
        return []

    processed = load_processed_files()
    new_pdfs = []

    for pdf_file in papers_path.glob("*.pdf"):
        if pdf_file.name not in processed:
            new_pdfs.append(str(pdf_file))

    return new_pdfs

def find_new_pdfs_from_drive() -> List[Tuple[str, str]]:
    """Find all new PDFs from Google Drive. Returns list of (file_id, file_name)."""
    drive_files = get_pdfs_from_drive()
    if not drive_files:
        logger.warning("No PDFs found in Google Drive folder")
        return []

    processed = load_processed_files()
    new_pdfs = []

    for file_id, file_name in drive_files:
        if file_name not in processed:
            new_pdfs.append((file_id, file_name))

    return new_pdfs

def process_single_pdf(pdf_path: str) -> bool:
    """Process a single local PDF file (for backward compatibility)."""
    pdf_name = os.path.basename(pdf_path)
    logger.info(f"Processing: {pdf_name}")

    # Extract text
    paper_text = extract_text_from_pdf(pdf_path)
    if not paper_text:
        logger.error(f"Could not extract text from {pdf_name}")
        return False

    logger.info(f"Extracted {len(paper_text)} characters from {pdf_name}")

    # Extract data locally (no API)
    extracted_data = extract_data_locally(paper_text, pdf_path)
    if not extracted_data:
        logger.error(f"Failed to extract data from {pdf_name}")
        return False

    # Write to Excel
    if not write_extraction_to_excel(extracted_data, pdf_name):
        logger.error(f"Failed to write extraction to Excel for {pdf_name}")
        return False

    # Mark as processed
    processed = load_processed_files()
    processed.add(pdf_name)
    save_processed_files(processed)

    logger.info(f"✓ Successfully processed {pdf_name}")
    return True

def process_single_pdf_from_drive(file_id: str, file_name: str) -> bool:
    """Process a single PDF from Google Drive without downloading."""
    logger.info(f"Processing: {file_name}")

    # Download PDF content from Google Drive
    pdf_bytes = download_pdf_from_drive(file_id)
    if not pdf_bytes:
        logger.error(f"Could not download {file_name} from Google Drive")
        return False

    # Extract text from bytes
    paper_text = extract_text_from_pdf_bytes(pdf_bytes, file_name)
    if not paper_text:
        logger.error(f"Could not extract text from {file_name}")
        return False

    logger.info(f"Extracted {len(paper_text)} characters from {file_name}")

    # Extract data locally (no API)
    extracted_data = extract_data_locally(paper_text, file_name)
    if not extracted_data:
        logger.error(f"Failed to extract data from {file_name}")
        return False

    # Write to Excel
    if not write_extraction_to_excel(extracted_data, file_name):
        logger.error(f"Failed to write extraction to Excel for {file_name}")
        return False

    # Mark as processed
    processed = load_processed_files()
    processed.add(file_name)
    save_processed_files(processed)

    logger.info(f"✓ Successfully processed {file_name}")
    return True


def create_excel_file_if_not_exists() -> bool:
    """Create Excel file with headers if it doesn't exist."""
    if os.path.exists(EXCEL_FILE):
        return True

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extractions"

        # Define headers (A-Y: 25 columns)
        headers = [
            "S.No.", "Author(s)", "Country/Location", "Title of Source",
            "Type of Publication", "Where Published", "Publication Year",
            "Type of Evidence Source", "Aim/Purpose", "Research Questions/Hypothesis",
            "Study Design", "Methodology and Methods", "Population and Sample Size",
            "Setting/Context", "Date and Location of Data Collection",
            "Definition of Coastal Areas", "Interventions/Programmes", "Outcomes",
            "Key Findings", "Knowledge Gaps", "Barriers, Facilitators, Limitations",
            "Institutional Affiliation", "Funder/Funding", "Data Accessibility",
            "Date & Time of Extraction"
        ]

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

        # Set column widths
        column_widths = {
            'A': 6, 'B': 20, 'C': 18, 'D': 25, 'E': 18, 'F': 18, 'G': 14, 'H': 18,
            'I': 18, 'J': 22, 'K': 18, 'L': 20, 'M': 20, 'N': 18, 'O': 22,
            'P': 18, 'Q': 20, 'R': 18, 'S': 20, 'T': 18, 'U': 25, 'V': 20, 'W': 18, 'X': 18, 'Y': 22
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(EXCEL_FILE)
        logger.info(f"✓ Created Excel file: {EXCEL_FILE}")
        return True
    except Exception as e:
        logger.error(f"Failed to create Excel file: {e}")
        return False

def run_extraction_agent() -> None:
    """Main workflow: get PDFs from Google Drive, process new ones."""
    logger.info("=" * 60)
    logger.info(f"Starting extraction agent at {datetime.now()}")
    logger.info("=" * 60)

    # Create Excel file if it doesn't exist
    if not create_excel_file_if_not_exists():
        logger.error("Failed to create/verify Excel file")
        return

    # Find new PDFs from Google Drive
    new_pdfs = find_new_pdfs_from_drive()

    if not new_pdfs:
        logger.info("No new PDFs found in Google Drive")
        return

    logger.info(f"Found {len(new_pdfs)} new PDF(s)")

    # Process each PDF from Google Drive
    successful = 0
    for file_id, file_name in new_pdfs:
        if process_single_pdf_from_drive(file_id, file_name):
            successful += 1

    logger.info("=" * 60)
    logger.info(f"Extraction complete: {successful}/{len(new_pdfs)} successful")
    logger.info("=" * 60)


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    run_extraction_agent()
